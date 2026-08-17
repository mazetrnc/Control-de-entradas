import tkinter as tk
from tkinter import messagebox, simpledialog
from datetime import datetime
import pandas as pd
from playsound3 import playsound
from openpyxl import load_workbook, Workbook
import os
import shutil
import logging
import atexit
import sys
import sqlite3
import smtplib
from email.message import EmailMessage

# El generador de reportes vive en su propio archivo (reporte_quincenal.py,
# debe estar guardado junto a este script) porque usa librerías adicionales
# (matplotlib, reportlab) que no hacen falta para el resto del programa. Si
# no están instaladas, el programa sigue funcionando con normalidad — solo
# no se generarán reportes automáticos (se avisa en el log).
try:
    from reporte_quincenal import generar_reporte_quincenal, identificar_quincena, rango_desde_identificador
    REPORTES_DISPONIBLES = True
except Exception:
    REPORTES_DISPONIBLES = False

# El reporte mensual detallado (por estudiante, para detectar candidatos a
# ceder su cupo de alimentación) vive en su propio archivo separado del
# quincenal, por la misma razón: reportlab sin matplotlib. Debe estar en la
# misma carpeta que este script.
try:
    from reporte_mensual_detallado import (
        generar_reporte_mensual_detallado,
        identificar_mes as identificar_mes_reporte,
        rango_desde_identificador_mes,
    )
    REPORTE_MENSUAL_DISPONIBLE = True
except Exception:
    REPORTE_MENSUAL_DISPONIBLE = False

# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================

# Carpeta donde vive la aplicación, sin importar desde dónde se ejecute
# (doble clic, acceso directo, otra terminal, etc.) y sin importar si corre
# como script de Python o como ejecutable compilado con PyInstaller.
#
# - Como script normal: __file__ apunta al .py real -> se usa esa carpeta.
# - Compilado con PyInstaller: sys.frozen es True y __file__ apunta a una
#   carpeta TEMPORAL que se borra al cerrar el programa (_MEIPASS), así que
#   ahí NUNCA deben guardarse datos; en su lugar se usa la carpeta donde
#   está el .exe (sys.executable).
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def ruta_recurso(nombre_archivo):
    """Ruta a un archivo de RECURSO INTERNO de la app (como el logo) --
    algo que se distribuye pegado al programa y nunca cambia el usuario,
    a diferencia de 'datos/' (que sí es del usuario y vive junto al .exe).
    Compilado con PyInstaller, estos recursos van empaquetados DENTRO del
    .exe y se extraen a una carpeta temporal (sys._MEIPASS) en cada
    arranque; como script normal, se busca en 'recursos/' junto al .py."""
    raiz = getattr(sys, '_MEIPASS', BASE_DIR)
    return os.path.join(raiz, 'recursos', nombre_archivo)

# Todos los archivos "de trabajo" (IDs, historiales diarios, historial
# total, contraseñas, log, bloqueo, sonido) viven ordenados dentro de
# 'datos/', separados de 'historicos/' (que guarda las copias mensuales).
DATA_DIR = os.path.join(BASE_DIR, 'datos')
CARPETA_HISTORICOS = os.path.join(BASE_DIR, 'historicos')
os.makedirs(DATA_DIR, exist_ok=True)

HISTORIAL_TOTAL_PATH = os.path.join(DATA_DIR, 'historial_total.xlsx')
ESTADO_MES_PATH = os.path.join(DATA_DIR, 'ultimo_mes.txt')
SONIDO_ERROR = os.path.join(DATA_DIR, 'Microsoft Windows 98 Error.mp3')
LOG_PATH = os.path.join(DATA_DIR, 'app.log')
LOCK_PATH = os.path.join(DATA_DIR, 'app.lock')
CIERRE_PATH = os.path.join(DATA_DIR, 'cierre.txt')
REIN_PATH = os.path.join(DATA_DIR, 'rein.txt')
SALIR_PANTALLA_PATH = os.path.join(DATA_DIR, 'salir_pantalla.txt')
ULTIMO_DIA_PATH = os.path.join(DATA_DIR, 'ultimo_dia.txt')
DB_PATH = os.path.join(DATA_DIR, 'asistencia.db')
ULTIMO_REPORTE_PATH = os.path.join(DATA_DIR, 'ultimo_reporte.txt')
ULTIMO_REPORTE_MENSUAL_PATH = os.path.join(DATA_DIR, 'ultimo_reporte_mensual.txt')
CARPETA_REPORTES_CONFIG_PATH = os.path.join(DATA_DIR, 'carpeta_reportes.txt')
CORREO_CONFIG_PATH = os.path.join(DATA_DIR, 'correo_config.txt')
DIRECTIVOS_CORREO_PATH = os.path.join(DATA_DIR, 'directivos_correo.txt')

# Cada categoría define su propio archivo de IDs válidos y su propio
# historial diario. Para agregar una categoría nueva en el futuro basta
# con añadir un diccionario aquí; el resto del programa se adapta solo.
CATEGORIAS = [
    {"nombre": "Estudiantes secundaria", "ids_path": os.path.join(DATA_DIR, "ids_secundaria.xlsx"), "historial_path": os.path.join(DATA_DIR, "vhistorial_secundaria.xlsx")},
    {"nombre": "Estudiantes primaria",   "ids_path": os.path.join(DATA_DIR, "ids_primaria.xlsx"),   "historial_path": os.path.join(DATA_DIR, "vhistorial_primaria.xlsx")},
    {"nombre": "Docentes",               "ids_path": os.path.join(DATA_DIR, "ids_docentes.xlsx"),  "historial_path": os.path.join(DATA_DIR, "vhistorial_docentes.xlsx")},
]

COLUMNAS_IDS = ['ID', 'Nombre', 'Apellido', 'Grado']
COLUMNAS_HISTORIAL = ['Fecha', 'Hora', 'ID', 'Nombre', 'Apellido', 'Grado']


# ============================================================
# LOG DE ERRORES / EVENTOS
# ============================================================
# Todo lo relevante (creación de archivos, sincronización, limpieza mensual,
# errores inesperados al escanear) queda registrado en 'app.log' con fecha y
# hora, para poder revisar después qué pasó un día que nadie estuvo presente.

_handlers = [logging.FileHandler(LOG_PATH, encoding='utf-8')]
if sys.stderr is not None:
    # Si se compila con PyInstaller en modo --windowed (sin consola),
    # sys.stderr es None y agregar un StreamHandler tronaría el logging.
    _handlers.append(logging.StreamHandler())

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=_handlers,
)


# ============================================================
# BLOQUEO SIMPLE PARA DETECTAR OTRA INSTANCIA ABIERTA
# ============================================================
# No impide abrir el programa (para no dejar a nadie fuera por un bloqueo
# viejo tras un cierre inesperado), pero deja aviso en el log/consola si ya
# había un 'app.lock', que suele significar que el programa sigue corriendo
# en otra parte (ej. otra computadora usando la misma carpeta compartida).

def adquirir_bloqueo():
    if os.path.exists(LOCK_PATH):
        logging.warning(
            "Se encontró 'app.lock' de una sesión anterior. Si el programa ya está "
            "abierto en otro equipo usando esta misma carpeta, cerrar uno de los dos "
            "evita que se pisen las escrituras del Excel."
        )
    try:
        with open(LOCK_PATH, "w") as f:
            f.write(datetime.now().isoformat())
    except Exception:
        logging.exception("No se pudo crear el archivo de bloqueo 'app.lock'")


def liberar_bloqueo():
    try:
        if os.path.exists(LOCK_PATH):
            os.remove(LOCK_PATH)
    except Exception:
        logging.exception("No se pudo eliminar el archivo de bloqueo 'app.lock'")


adquirir_bloqueo()
atexit.register(liberar_bloqueo)


def reproducir_sonido_error():
    """Envuelto en try/except: si el archivo de sonido falta o el equipo no
    tiene salida de audio configurada, no debe tumbar la verificación."""
    try:
        playsound(SONIDO_ERROR, block=False)
    except Exception:
        logging.warning("No se pudo reproducir el sonido de error.", exc_info=True)


# ============================================================
# UTILIDADES DE ARCHIVOS (crean lo que falte, nunca truenan)
# ============================================================

def asegurar_archivo_ids(ruta, categoria_nombre=""):
    """Si el archivo de IDs de una categoría no existe (por ejemplo
    'ids_docentes.xlsx' la primera vez que se usa este programa), se crea
    una plantilla con UNA hoja de ejemplo. El formato es: una hoja POR
    GRADO/GRUPO (el nombre de la hoja ES el grado), cada una con columnas
    ID, Nombre, Apellido — igual que las hojas de historial_total.xlsx, así
    es más fácil para quien captura buscar/organizar por grado."""
    if not os.path.exists(ruta):
        wb = Workbook()
        hoja = wb.active
        hoja.title = "EJEMPLO_Grado"
        hoja.append(["ID", "Nombre", "Apellido"])
        wb.save(ruta)
        logging.info(
            f"Se creó la plantilla '{ruta}' para {categoria_nombre}. Renombra/duplica la hoja "
            f"'EJEMPLO_Grado' — crea una hoja por cada grado o grupo (el NOMBRE de la hoja es "
            f"el grado) con columnas ID, Nombre, Apellido."
        )


def leer_ids_categoria(ruta):
    """Lee un archivo de IDs organizado en varias hojas (una por grado o
    grupo, igual que historial_total.xlsx): cada hoja debe tener columnas
    ID, Nombre, Apellido, y el NOMBRE de la hoja se usa como 'Grado'.

    Devuelve un único DataFrame con columnas ID, Nombre, Apellido, Grado,
    juntando todas las hojas que tengan datos (las vacías se ignoran).
    También admite, por compatibilidad, que una hoja ya traiga su propia
    columna 'Grado' llena (en ese caso se respeta en vez de usar el nombre
    de la hoja)."""
    try:
        hojas = pd.read_excel(ruta, sheet_name=None, dtype=str)
    except Exception:
        logging.exception(f"No se pudo leer el archivo de IDs '{ruta}'")
        return pd.DataFrame(columns=COLUMNAS_IDS)

    partes = []
    for nombre_hoja, df in hojas.items():
        if df is None or df.empty:
            continue
        df = df.copy()
        if 'Grado' not in df.columns or df['Grado'].isna().all():
            df['Grado'] = nombre_hoja
        else:
            df['Grado'] = df['Grado'].fillna(nombre_hoja)
        for col in COLUMNAS_IDS:
            if col not in df.columns:
                df[col] = None
        partes.append(df[COLUMNAS_IDS])

    if not partes:
        return pd.DataFrame(columns=COLUMNAS_IDS)

    todos = pd.concat(partes, ignore_index=True)
    todos = todos.dropna(subset=['ID'])
    todos['ID'] = todos['ID'].astype(str).str.strip()
    todos = todos[todos['ID'] != ""]
    return todos


def asegurar_historial_diario(ruta):
    if not os.path.exists(ruta):
        pd.DataFrame(columns=COLUMNAS_HISTORIAL).to_excel(ruta, index=False)


def asegurar_archivo_password(ruta, valor_por_defecto):
    """Crea un archivo de contraseña simple (una por línea) si no existe,
    para que el programa no truene la primera vez que se usa esta función."""
    if not os.path.exists(ruta):
        with open(ruta, "w") as f:
            f.write(valor_por_defecto + "\n")
        logging.warning(
            f"Se creó '{ruta}' con una contraseña por defecto ('{valor_por_defecto}'). "
            f"Cámbiala por una propia antes de usar el programa en producción."
        )


def asegurar_archivo_carpeta_reportes_config():
    """Crea la plantilla de 'carpeta_reportes.txt' si no existe. Vacío (o
    solo con comentarios) por default: en ese caso los reportes solo se
    guardan en la carpeta local 'reportes/'. Si se escribe ahí la ruta de
    una carpeta sincronizada (Google Drive, OneDrive, etc.), los reportes
    se guardan TAMBIÉN ahí (además de la copia local, nunca en vez de
    ella) y esa misma app de sincronización los sube sola a la nube — sin
    que el operador tenga que hacer nada más.

    IMPORTANTE: la copia local en 'reportes/' siempre se genera primero y
    siempre queda, pase lo que pase con la carpeta de nube (permisos,
    carpeta desconectada, app de sincronización caída, etc.) — es el
    respaldo garantizado."""
    if not os.path.exists(CARPETA_REPORTES_CONFIG_PATH):
        contenido = (
            "# Escribe aquí (sin el #) la ruta completa de la carpeta donde quieres\n"
            "# que se guarde una copia ADICIONAL de los reportes (quincenales y\n"
            "# mensuales) -- por ejemplo, tu carpeta sincronizada de Google Drive o\n"
            "# OneDrive. Los reportes SIEMPRE se guardan primero en la carpeta local\n"
            "# 'reportes', junto al programa (ese respaldo local nunca se salta,\n"
            "# pase lo que pase con esta carpeta de nube). Si dejas este archivo\n"
            "# vacío o solo con comentarios, la copia local es la única.\n"
            "#\n"
            "# Ejemplo en Windows:\n"
            "# C:\\Users\\TuUsuario\\Google Drive\\Reportes PAE\n"
        )
        with open(CARPETA_REPORTES_CONFIG_PATH, "w", encoding="utf-8") as f:
            f.write(contenido)


def obtener_carpeta_reportes_nube():
    """Lee 'carpeta_reportes.txt' y devuelve la primera línea que no sea un
    comentario (#) ni esté vacía, como ruta de una carpeta de nube donde
    copiar ADICIONALMENTE los reportes. Si el archivo no existe, está
    vacío, o la línea está en blanco, devuelve None (no hay carpeta de
    nube configurada; la copia local sigue generándose de todas formas)."""
    if os.path.exists(CARPETA_REPORTES_CONFIG_PATH):
        with open(CARPETA_REPORTES_CONFIG_PATH, "r", encoding="utf-8-sig") as f:
            for linea in f:
                linea = linea.strip()
                if linea and not linea.startswith("#"):
                    return linea
    return None


def asegurar_archivo_config_correo():
    """Crea la plantilla de 'correo_config.txt' si no existe. Mientras no se
    complete, el envío automático por correo simplemente no ocurre (el
    reporte de todas formas queda guardado en la carpeta de reportes)."""
    if not os.path.exists(CORREO_CONFIG_PATH):
        contenido = (
            "# Configuración para enviar el reporte quincenal por correo automáticamente.\n"
            "# Completa los 5 valores de abajo (quita el # de cada línea que uses).\n"
            "# Mientras falte alguno de servidor/puerto/remitente/contraseña_app, el\n"
            "# correo simplemente no se envía (sin errores). 'seguridad' es opcional,\n"
            "# si no se indica se usa 'ssl' por default.\n"
            "#\n"
            "# ---- Opción 1: Gmail (gratis, necesita verificación por teléfono) ----\n"
            "# Crea una 'contraseña de aplicación' en la configuración de seguridad\n"
            "# de tu cuenta de Google (NO uses la contraseña normal de la cuenta):\n"
            "#\n"
            "# servidor=smtp.gmail.com\n"
            "# puerto=465\n"
            "# remitente=tu_correo@gmail.com\n"
            "# contraseña_app=xxxxxxxxxxxxxxxx\n"
            "# seguridad=ssl\n"
            "#\n"
            "# ---- Opción 2: Proton Mail (requiere plan de pago Mail Plus o superior) ----\n"
            "# Necesitas instalar y tener ABIERTA la app 'Proton Mail Bridge' en esta\n"
            "# misma computadora (actívale 'iniciar con el sistema' en su configuración).\n"
            "# Dentro de Bridge verás el servidor, puerto y una contraseña generada\n"
            "# especial para esto (NO es la contraseña de tu cuenta Proton). El puerto\n"
            "# más común es 1025, pero Bridge puede asignar otro si ese ya está\n"
            "# ocupado -- usa el que Bridge te muestre, no lo adivines:\n"
            "#\n"
            "# servidor=127.0.0.1\n"
            "# puerto=1025\n"
            "# remitente=tu_correo@proton.me\n"
            "# contraseña_app=la_contraseña_que_te_dio_bridge\n"
            "# seguridad=starttls\n"
        )
        with open(CORREO_CONFIG_PATH, "w", encoding="utf-8") as f:
            f.write(contenido)


def asegurar_archivo_directivos_correo():
    """Crea la plantilla de 'directivos_correo.txt' si no existe."""
    if not os.path.exists(DIRECTIVOS_CORREO_PATH):
        contenido = (
            "# Un correo de directivo por línea (sin el #). Se les manda una copia\n"
            "# del reporte automáticamente en cuanto se genera cada quincena.\n"
            "#\n"
            "# ejemplo_directivo@correo.com\n"
        )
        with open(DIRECTIVOS_CORREO_PATH, "w", encoding="utf-8") as f:
            f.write(contenido)


def leer_config_correo():
    config = {}
    if os.path.exists(CORREO_CONFIG_PATH):
        with open(CORREO_CONFIG_PATH, "r", encoding="utf-8-sig") as f:
            for linea in f:
                linea = linea.strip()
                if not linea or linea.startswith("#") or "=" not in linea:
                    continue
                clave, _, valor = linea.partition("=")
                config[clave.strip().lower()] = valor.strip()
    # Google muestra la contraseña de aplicación en grupos de 4 separados por
    # espacios (solo para que se lea más fácil) -- si se copió tal cual con
    # esos espacios, se quitan aquí para que no cause fallos de autenticación.
    if "contraseña_app" in config:
        config["contraseña_app"] = config["contraseña_app"].replace(" ", "")
    return config


def leer_directivos_correo():
    destinatarios = []
    if os.path.exists(DIRECTIVOS_CORREO_PATH):
        with open(DIRECTIVOS_CORREO_PATH, "r", encoding="utf-8-sig") as f:
            for linea in f:
                linea = linea.strip()
                if linea and not linea.startswith("#"):
                    destinatarios.append(linea)
    return destinatarios


def distribuir_reporte(ruta_pdf_local):
    """El PDF ya se generó en CARPETA_REPORTES_LOCAL (el respaldo garantizado
    -- eso siempre sucede primero, sin importar esta función). Si además hay
    una carpeta de nube configurada en 'carpeta_reportes.txt', se copia el
    mismo archivo ahí también. Si esa copia falla (carpeta sin permisos,
    ruta que ya no existe, app de sincronización caída, etc.) NUNCA se
    pierde el reporte: la copia local ya quedó guardada de todas formas, y
    aquí solo se anota el problema en el log."""
    if not CARPETA_REPORTES_NUBE:
        return  # no hay carpeta de nube configurada; la copia local es la única y ya está lista
    try:
        os.makedirs(CARPETA_REPORTES_NUBE, exist_ok=True)
        destino = os.path.join(CARPETA_REPORTES_NUBE, os.path.basename(ruta_pdf_local))
        shutil.copy(ruta_pdf_local, destino)
        logging.info(f"Copia adicional del reporte guardada en la carpeta de nube configurada: {destino}")
    except Exception:
        logging.exception(
            "No se pudo copiar el reporte a la carpeta de nube configurada en 'carpeta_reportes.txt'. "
            "El reporte de todas formas quedó guardado con normalidad en la carpeta local 'reportes/'."
        )


def enviar_reporte_por_correo(ruta_pdf, etiqueta_periodo, tipo="quincenal"):
    """Envía el PDF del reporte por correo a los directivos configurados en
    'directivos_correo.txt', usando las credenciales de 'correo_config.txt'.
    Si falta completar cualquiera de los dos archivos, no envía nada (el
    reporte de todas formas ya quedó guardado normalmente) y solo lo anota
    en el log — nunca truena el programa.

    'tipo' es "quincenal" o "mensual", solo para ajustar el texto del
    asunto y el cuerpo del correo.

    Soporta dos modos de conexión, según la clave 'seguridad' en
    correo_config.txt:
      - 'ssl' (default): conexión SSL directa. Es lo que usa Gmail (puerto 465).
      - 'starttls': conexión que empieza sin cifrar y luego la "sube" a
        segura. Es lo que usa Proton Mail Bridge (servidor 127.0.0.1,
        puerto 1025 típicamente) y muchos otros proveedores (puerto 587).
    """
    config = leer_config_correo()
    destinatarios = leer_directivos_correo()

    faltantes = [c for c in ("servidor", "puerto", "remitente", "contraseña_app") if not config.get(c)]
    if faltantes or not destinatarios:
        detalle = f"faltan: {', '.join(faltantes)}" if faltantes else "no hay destinatarios en directivos_correo.txt"
        logging.info(f"No se envió el reporte por correo todavía ({detalle}). El PDF sigue guardado con normalidad.")
        return

    if tipo == "mensual":
        asunto = f"Reporte mensual detallado de asistencia PAE — {etiqueta_periodo}"
        cuerpo = (
            f"Buen día,\n\nSe adjunta el reporte mensual detallado de asistencia (por estudiante/docente) "
            f"correspondiente a {etiqueta_periodo}. Incluye el % de asistencia individual y quiénes están "
            f"por debajo del umbral definido como candidatos a ceder su cupo de alimentación.\n\n"
            f"Este correo se generó automáticamente, no es necesario responderlo."
        )
    else:
        asunto = f"Reporte de asistencia PAE — {etiqueta_periodo}"
        cuerpo = (
            f"Buen día,\n\nSe adjunta el reporte de asistencia correspondiente a la quincena "
            f"{etiqueta_periodo}.\n\nEste correo se generó automáticamente, no es necesario responderlo."
        )

    try:
        mensaje = EmailMessage()
        mensaje["Subject"] = asunto
        mensaje["From"] = config["remitente"]
        mensaje["To"] = ", ".join(destinatarios)
        mensaje.set_content(cuerpo)
        with open(ruta_pdf, "rb") as f:
            mensaje.add_attachment(
                f.read(), maintype="application", subtype="pdf", filename=os.path.basename(ruta_pdf)
            )

        modo_seguridad = config.get("seguridad", "ssl").strip().lower()
        if modo_seguridad == "starttls":
            with smtplib.SMTP(config["servidor"], int(config["puerto"])) as smtp:
                smtp.starttls()
                smtp.login(config["remitente"], config["contraseña_app"])
                smtp.send_message(mensaje)
        else:
            with smtplib.SMTP_SSL(config["servidor"], int(config["puerto"])) as smtp:
                smtp.login(config["remitente"], config["contraseña_app"])
                smtp.send_message(mensaje)

        logging.info(f"Reporte ({tipo}) enviado por correo a: {', '.join(destinatarios)}")
    except Exception:
        logging.exception("No se pudo enviar el reporte por correo (el PDF sigue guardado localmente con normalidad)")


for _cat in CATEGORIAS:
    asegurar_archivo_ids(_cat["ids_path"], _cat["nombre"])
    asegurar_historial_diario(_cat["historial_path"])

asegurar_archivo_carpeta_reportes_config()
asegurar_archivo_config_correo()
asegurar_archivo_directivos_correo()

# Carpeta LOCAL donde SIEMPRE se genera primero cada reporte (quincenal o
# mensual) -- el respaldo garantizado, sin importar la configuración de
# nube. CARPETA_REPORTES_NUBE es opcional y, si está configurada, recibe
# una COPIA adicional (ver distribuir_reporte más abajo).
CARPETA_REPORTES_LOCAL = os.path.join(BASE_DIR, 'reportes')
CARPETA_REPORTES_NUBE = obtener_carpeta_reportes_nube()

asegurar_archivo_password(SALIR_PANTALLA_PATH, "0000")


# ============================================================
# BASE DE DATOS SQLite (FASE 1: en paralelo al Excel)
# ============================================================
# Todo lo que ya se guarda en Excel se sigue guardando exactamente igual
# que antes — esto SOLO agrega una copia adicional, más fácil de analizar,
# en un único archivo 'datos/asistencia.db'. Si algo de esto falla, nunca
# debe impedir que el programa siga funcionando con Excel como hasta hoy;
# por eso cada función de esta sección va envuelta en su propio try/except.

_conexion_db = None


def obtener_conexion_db():
    """Devuelve la conexión al archivo .db, reutilizando la misma durante
    toda la sesión en vez de abrir y cerrar una nueva en cada operación
    (eso era lo que hacía más lento cada escaneo: abrir/cerrar el archivo
    y forzar una escritura sincronizada a disco en cada commit).

    Al abrir la conexión por primera vez, se activa el modo WAL
    ('Write-Ahead Logging'), la configuración recomendada por SQLite
    justamente para este patrón de uso (una sola aplicación, muchas
    escrituras pequeñas y frecuentes) — sigue siendo seguro ante un
    apagón o cierre inesperado, pero mucho más rápido que el modo por
    default."""
    global _conexion_db
    if _conexion_db is None:
        _conexion_db = sqlite3.connect(DB_PATH)
        _conexion_db.execute("PRAGMA journal_mode=WAL")
        _conexion_db.execute("PRAGMA synchronous=NORMAL")
    return _conexion_db


def cerrar_conexion_db():
    global _conexion_db
    if _conexion_db is not None:
        try:
            _conexion_db.close()
        except Exception:
            logging.exception("No se pudo cerrar la conexión de asistencia.db limpiamente")
        _conexion_db = None


atexit.register(cerrar_conexion_db)


def inicializar_base_datos():
    """Crea las 3 tablas si no existen todavía (no borra nada si ya
    existen). Es el equivalente a poner los encabezados de una hoja nueva:

    - personas: una fila por estudiante/docente (como los ids_*.xlsx, pero
      los tres archivos juntos en un solo lugar).
    - entradas: una fila por CADA vez que alguien registra su entrada
      (esto reemplaza las "X" que hoy se marcan en historial_total.xlsx).
    - estadisticas_diarias: un resumen por día y por categoría (reemplaza
      la hoja "Estadisticas"). Se usa (fecha, categoria) como identificador
      único, así que sirve para cualquier cantidad de categorías, no solo
      las 3 actuales.
    """
    conexion = obtener_conexion_db()
    conexion.execute("""
        CREATE TABLE IF NOT EXISTS personas (
            id         TEXT PRIMARY KEY,
            nombre     TEXT,
            apellido   TEXT,
            grado      TEXT,
            categoria  TEXT
        )
    """)
    conexion.execute("""
        CREATE TABLE IF NOT EXISTS entradas (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            persona_id  TEXT,
            fecha       TEXT,
            hora        TEXT,
            FOREIGN KEY (persona_id) REFERENCES personas(id)
        )
    """)
    conexion.execute("""
        CREATE TABLE IF NOT EXISTS estadisticas_diarias (
            fecha       TEXT,
            categoria   TEXT,
            porcentaje  REAL,
            PRIMARY KEY (fecha, categoria)
        )
    """)
    conexion.commit()


def sync_personas_db():
    """Sincroniza la tabla 'personas' con los tres archivos ids_*.xlsx —
    el mismo trabajo que ya hace sync_historial_total() para el Excel,
    pero aquí. 'INSERT OR REPLACE' significa: si el ID ya existe, actualiza
    su nombre/apellido/grado; si no existe, lo agrega. Se llama cada vez
    que arranca el programa, igual que la sincronización del Excel."""
    conexion = obtener_conexion_db()
    for cat in CATEGORIAS:
        df = leer_ids_categoria(cat["ids_path"])
        for _, persona in df.iterrows():
            conexion.execute(
                """
                INSERT OR REPLACE INTO personas (id, nombre, apellido, grado, categoria)
                VALUES (?, ?, ?, ?, ?)
                """,
                (persona["ID"], persona["Nombre"], persona["Apellido"], persona["Grado"], cat["nombre"]),
            )
    conexion.commit()


def registrar_entrada_db(persona_id, fecha, hora):
    """Agrega UNA fila a la tabla 'entradas'. Se llama cada vez que se
    registra una entrada válida, justo junto a donde se escribe en el
    Excel diario — así las dos copias siempre quedan sincronizadas."""
    try:
        conexion = obtener_conexion_db()
        conexion.execute(
            "INSERT INTO entradas (persona_id, fecha, hora) VALUES (?, ?, ?)",
            (persona_id, fecha, hora),
        )
        conexion.commit()
    except Exception:
        logging.exception(f"No se pudo registrar en asistencia.db la entrada de '{persona_id}'")


def registrar_estadisticas_db(fecha, porcentajes_por_categoria):
    """Guarda (o actualiza) el % de asistencia de un día para cada
    categoría. 'porcentajes_por_categoria' es un diccionario como
    {'Estudiantes secundaria': 0.85, 'Estudiantes primaria': 0.6, ...}."""
    try:
        conexion = obtener_conexion_db()
        for categoria, porcentaje in porcentajes_por_categoria.items():
            conexion.execute(
                """
                INSERT INTO estadisticas_diarias (fecha, categoria, porcentaje)
                VALUES (?, ?, ?)
                ON CONFLICT(fecha, categoria) DO UPDATE SET porcentaje = excluded.porcentaje
                """,
                (fecha, categoria, porcentaje),
            )
        conexion.commit()
    except Exception:
        logging.exception(f"No se pudieron registrar en asistencia.db las estadísticas del {fecha}")


try:
    inicializar_base_datos()
    sync_personas_db()
except Exception:
    logging.exception("Error al inicializar/sincronizar asistencia.db (el programa sigue con Excel con normalidad)")


# ============================================================
# SINCRONIZACIÓN AUTOMÁTICA DE historial_total.xlsx
# ============================================================

def crear_hoja_grado(wb, grado):
    """Crea una hoja nueva con la misma estructura que usa add_to_level_sheet:
    fila 1 = fechas de asistencia, fila 2 = encabezados, fila 3 en adelante = personas."""
    hoja = wb.create_sheet(title=str(grado))
    hoja.cell(row=2, column=1, value="Nombre")
    hoja.cell(row=2, column=2, value="Apellido")
    hoja.cell(row=2, column=3, value="ID")
    return hoja


def sync_historial_total():
    """Mantiene historial_total.xlsx sincronizado con los archivos de IDs de
    TODAS las categorías (primaria, secundaria, docentes, y cualquiera que se
    agregue después):
      - Crea el archivo si no existe.
      - Crea una hoja nueva por cada grado/grupo que aparezca en los IDs.
      - Agrega las filas de las personas nuevas que todavía no estén.
      - Actualiza nombre/apellido si cambiaron en el Excel de IDs.
      - Avisa en el log si el mismo ID aparece repetido entre categorías.
    Nunca borra ni toca la asistencia que ya se había marcado.
    """
    dfs = []
    for cat in CATEGORIAS:
        df = leer_ids_categoria(cat["ids_path"])
        if not df.empty:
            dfs.append(df)
    if not dfs:
        return

    todos = pd.concat(dfs, ignore_index=True)
    todos = todos.dropna(subset=["ID"])
    todos["ID"] = todos["ID"].astype(str).str.strip()

    duplicados = todos[todos.duplicated(subset=["ID"], keep=False)]
    if not duplicados.empty:
        ids_dup = sorted(duplicados["ID"].unique())
        logging.warning(
            f"Se encontraron IDs duplicados entre los archivos de IDs (aparecen en más de "
            f"una categoría o más de una vez): {ids_dup}. Se usará el último registro "
            f"encontrado para cada uno; revisa esos archivos para corregirlo."
        )

    todos = todos.drop_duplicates(subset=["ID"], keep="last")

    if not os.path.exists(HISTORIAL_TOTAL_PATH):
        wb = Workbook()
        wb.remove(wb.active)  # quitar la hoja "Sheet" por defecto
    else:
        wb = load_workbook(HISTORIAL_TOTAL_PATH)

    grados = sorted(todos["Grado"].dropna().unique())

    for grado in grados:
        grado = str(grado)
        hoja = wb[grado] if grado in wb.sheetnames else crear_hoja_grado(wb, grado)

        # Personas que ya están en la hoja (columna 3, desde la fila 3)
        ids_en_hoja = {}
        for fila in range(3, hoja.max_row + 1):
            valor_id = hoja.cell(row=fila, column=3).value
            if valor_id is not None:
                ids_en_hoja[str(valor_id).strip()] = fila

        df_grado = todos[todos["Grado"].astype(str) == grado]
        siguiente_fila = max(hoja.max_row, 2) + 1

        for _, persona in df_grado.iterrows():
            pid = str(persona["ID"]).strip()
            nombre = persona.get("Nombre", "")
            apellido = persona.get("Apellido", "")
            if pid in ids_en_hoja:
                fila = ids_en_hoja[pid]
                hoja.cell(row=fila, column=1, value=nombre)
                hoja.cell(row=fila, column=2, value=apellido)
            else:
                hoja.cell(row=siguiente_fila, column=1, value=nombre)
                hoja.cell(row=siguiente_fila, column=2, value=apellido)
                hoja.cell(row=siguiente_fila, column=3, value=pid)
                ids_en_hoja[pid] = siguiente_fila
                siguiente_fila += 1

    if "Estadisticas" not in wb.sheetnames:
        wb.create_sheet("Estadisticas")

    wb.save(HISTORIAL_TOTAL_PATH)


# ============================================================
# LIMPIEZA Y ARCHIVADO MENSUAL DE historial_total.xlsx
# ============================================================

def gestionar_limpieza_mensual():
    """Cada vez que se detecta que cambió el mes (comparando contra la última
    vez que se revisó, guardada en 'ultimo_mes.txt'):
      1. Exporta una COPIA completa de historial_total.xlsx a 'historicos/'
         con el nombre del mes que acaba de terminar (para hacer estudios
         estadísticos después).
      2. Limpia las marcas de asistencia del mes que terminó (conserva la
         lista de estudiantes/docentes, solo borra las columnas de fechas).
      3. Actualiza el registro del mes activo.
    """
    mes_actual = datetime.now().strftime("%Y-%m")

    if not os.path.exists(ESTADO_MES_PATH):
        with open(ESTADO_MES_PATH, "w") as f:
            f.write(mes_actual)
        return  # primera vez que corre el programa, todavía no hay nada que archivar

    with open(ESTADO_MES_PATH, "r", encoding="utf-8-sig") as f:
        mes_guardado = f.read().strip()

    if mes_guardado == mes_actual or not os.path.exists(HISTORIAL_TOTAL_PATH):
        if mes_guardado != mes_actual:
            with open(ESTADO_MES_PATH, "w") as f:
                f.write(mes_actual)
        return

    # 1. Copia de respaldo del mes que terminó (dentro de la subcarpeta 'historicos/')
    os.makedirs(CARPETA_HISTORICOS, exist_ok=True)
    archivo_respaldo = os.path.join(CARPETA_HISTORICOS, f"historial_total_{mes_guardado}.xlsx")
    if not os.path.exists(archivo_respaldo):
        shutil.copy(HISTORIAL_TOTAL_PATH, archivo_respaldo)
        logging.info(f"Historial del mes {mes_guardado} exportado como '{archivo_respaldo}'.")

    # 2. Limpiar asistencia (columnas de fecha, desde la 4), sin tocar "Estadisticas"
    wb = load_workbook(HISTORIAL_TOTAL_PATH)
    for sheet_name in wb.sheetnames:
        if sheet_name == "Estadisticas":
            continue
        hoja = wb[sheet_name]
        for col in range(4, hoja.max_column + 1):
            for row in range(1, hoja.max_row + 1):
                hoja.cell(row=row, column=col).value = None
    wb.save(HISTORIAL_TOTAL_PATH)

    # 3. Actualizar el mes de referencia
    with open(ESTADO_MES_PATH, "w") as f:
        f.write(mes_actual)

    logging.info(f"Limpieza mensual realizada. Nuevo mes activo: {mes_actual}")


# ============================================================
# REINICIO DIARIO DEL CONTEO (automático, con respaldo manual)
# ============================================================
# Antes, el único disparador era escanear el código de rein.txt. El
# problema: si un día se olvida escanearlo, el historial de "quién ya
# entró" nunca se limpia, y al día siguiente el sistema bloquearía a TODO
# el mundo como "ya ha entrado hoy". Por eso ahora el reinicio ocurre solo
# en cuanto cambia la fecha (comparando contra 'ultimo_dia.txt'), y el
# código manual de rein.txt se conserva como respaldo (por ejemplo para
# reiniciar a media jornada, o corregir un error de captura).

def ejecutar_reinicio_diario(motivo="automático", fecha=None):
    """Guarda en 'Estadisticas' el % de asistencia del día por categoría y
    limpia los historiales diarios de todas las categorías. La usan tanto
    el reinicio automático por cambio de fecha como el código manual de
    rein.txt, para no duplicar la lógica en dos lugares distintos.

    'fecha' es la fecha A LA QUE PERTENECE la asistencia que se está
    cerrando (formato 'YYYY-MM-DD'). Si no se indica, se usa la fecha de
    hoy — correcto para el reinicio manual (se escanea el mismo día), pero
    el reinicio automático SIEMPRE debe pasar explícitamente el día que
    acaba de terminar, no la fecha en que corre el reinicio (que ya es el
    día siguiente)."""
    if fecha is None:
        fecha = datetime.now().strftime("%Y-%m-%d")
    semana = datetime.strptime(fecha, "%Y-%m-%d").strftime("%V")
    resumen = {"Semana": [int(semana)], "Fecha": [fecha]}
    porcentajes_por_categoria = {}
    total_entradas_todas_categorias = 0

    for cat in CATEGORIAS:
        ids_cat = leer_ids_categoria(cat["ids_path"])
        total_cat = ids_cat['ID'].count()
        hist_cat = pd.read_excel(cat["historial_path"], dtype={'ID': str})
        entraron_cat = hist_cat['ID'].count()
        porcentaje = entraron_cat / total_cat if total_cat else 0
        resumen[cat["nombre"]] = [porcentaje]
        porcentajes_por_categoria[cat["nombre"]] = porcentaje
        total_entradas_todas_categorias += entraron_cat

        historial_vacio = pd.DataFrame(columns=COLUMNAS_HISTORIAL)
        historial_vacio.to_excel(cat["historial_path"], index=False)

    if total_entradas_todas_categorias == 0:
        # Nadie de NINGUNA categoría registró entrada ese día — lo más
        # probable es que haya sido festivo o un día sin clases/PAE, no que
        # el 100% de la comunidad haya faltado. Si se guardara este día con
        # 0% de asistencia, arrastraría hacia abajo los promedios (por
        # ejemplo "asistencia promedio los martes") de forma artificial.
        # Por eso NO se guarda una fila de estadísticas para esta fecha, ni
        # en la hoja "Estadisticas" ni en asistencia.db.
        logging.info(
            f"Reinicio diario ({motivo}): 0 entradas en TODAS las categorías el {fecha}. "
            f"Se asume día sin clases/PAE (festivo) y no se guardan estadísticas de esta fecha."
        )
    else:
        new_df = pd.DataFrame(resumen)
        save_statistics(new_df)
        registrar_estadisticas_db(fecha, porcentajes_por_categoria)

    logging.info(f"Reinicio diario ejecutado ({motivo}). Fecha registrada: {fecha}")


def gestionar_reinicio_diario_automatico():
    """Si cambió la fecha desde la última revisión (guardada en
    'ultimo_dia.txt'), reinicia solo los historiales diarios (no toca
    historial_total.xlsx, eso lo maneja la limpieza mensual). No depende de
    que nadie escanee nada: corre sola al abrir el programa cada mañana, y
    también se revisa en cada escaneo por si la aplicación se queda abierta
    de un día para otro. Devuelve True si acaba de ejecutar un reinicio.

    IMPORTANTE: la asistencia acumulada en los historiales diarios en este
    momento pertenece al último día que estuvo activo ('dia_guardado'), NO
    a 'hoy' (hoy es el día que apenas está empezando). Por eso se le pasa
    'dia_guardado' como fecha a ejecutar_reinicio_diario — si se usara
    'hoy' por error, las estadísticas quedarían fechadas un día después de
    cuando realmente ocurrió esa asistencia (y en el mes equivocado, si el
    cambio de fecha también cruzó de mes).
    """
    hoy = datetime.now().strftime("%Y-%m-%d")

    if not os.path.exists(ULTIMO_DIA_PATH):
        with open(ULTIMO_DIA_PATH, "w") as f:
            f.write(hoy)
        return False  # primera vez que corre el programa, nada que reiniciar todavía

    with open(ULTIMO_DIA_PATH, "r", encoding="utf-8-sig") as f:
        dia_guardado = f.read().strip()

    if dia_guardado == hoy:
        return False

    ejecutar_reinicio_diario(
        motivo=f"cambio automático de fecha ({dia_guardado} -> {hoy})",
        fecha=dia_guardado,
    )

    with open(ULTIMO_DIA_PATH, "w") as f:
        f.write(hoy)

    return True


# ============================================================
# REPORTE QUINCENAL AUTOMÁTICO
# ============================================================

def gestionar_reporte_quincenal():
    """Genera automáticamente el PDF del reporte de asistencia cada vez que
    se cierra una QUINCENA DE CALENDARIO — días 1 al 15, y del 16 al fin de
    mes — no cada 15 días corridos desde que se instaló el programa. Así
    los reportes siempre corresponden a "primera/segunda quincena de
    [mes]", igual que la entrega de recursos, en vez de ir desfasándose.

    Si no hay suficientes días con datos todavía dentro de esa quincena,
    reporte_quincenal.py simplemente no genera nada y se vuelve a intentar
    en la siguiente — nunca truena el programa por falta de datos."""
    if not REPORTES_DISPONIBLES:
        return

    hoy = datetime.now().date()
    id_quincena_actual, _, _ = identificar_quincena(hoy)

    if not os.path.exists(ULTIMO_REPORTE_PATH):
        # Primera vez que corre el programa: se guarda la quincena ACTUAL
        # como punto de partida, sin generar nada todavía (apenas estamos
        # a la mitad de esa quincena, casi no habría datos). El primer
        # reporte real saldrá en cuanto esta quincena termine.
        with open(ULTIMO_REPORTE_PATH, "w") as f:
            f.write(id_quincena_actual)
        return

    with open(ULTIMO_REPORTE_PATH, "r", encoding="utf-8-sig") as f:
        id_quincena_guardada = f.read().strip()

    if id_quincena_guardada == id_quincena_actual:
        return  # seguimos dentro de la misma quincena, nada que generar todavía

    # Cambiamos de quincena: se genera el reporte de la que ACABA DE
    # TERMINAR (id_quincena_guardada), no de la que apenas empieza hoy.
    try:
        inicio_anterior, fin_anterior = rango_desde_identificador(id_quincena_guardada)
        etiqueta_periodo = f"{inicio_anterior.strftime('%d %b %Y')} al {fin_anterior.strftime('%d %b %Y')}"
        # Siempre se genera primero en la carpeta LOCAL (respaldo garantizado).
        ruta_generada = generar_reporte_quincenal(
            DB_PATH, CARPETA_REPORTES_LOCAL, (inicio_anterior, fin_anterior), etiqueta_periodo
        )
        if ruta_generada:
            logging.info(f"Reporte quincenal generado ({etiqueta_periodo}): {ruta_generada}")
            distribuir_reporte(ruta_generada)
            enviar_reporte_por_correo(ruta_generada, etiqueta_periodo, tipo="quincenal")
        with open(ULTIMO_REPORTE_PATH, "w") as f:
            f.write(id_quincena_actual)
    except Exception:
        logging.exception("Error al generar el reporte quincenal")


# ============================================================
# REPORTE MENSUAL DETALLADO AUTOMÁTICO (por estudiante/docente)
# ============================================================

def gestionar_reporte_mensual():
    """Genera automáticamente el PDF del reporte MENSUAL DETALLADO (una fila
    por cada persona, con su % de asistencia del mes y quiénes son
    candidatos a ceder su cupo de alimentación) cada vez que se detecta que
    terminó un mes de calendario. Usa su propio archivo de seguimiento
    ('ultimo_reporte_mensual.txt'), independiente del que usa el reporte
    quincenal, así que uno no interfiere con el otro.

    Igual que el quincenal: si no hay suficientes días con datos todavía,
    reporte_mensual_detallado.py simplemente no genera nada y se vuelve a
    intentar el mes siguiente — nunca truena el programa."""
    if not REPORTE_MENSUAL_DISPONIBLE:
        return

    hoy = datetime.now().date()
    id_mes_actual, _, _ = identificar_mes_reporte(hoy)

    if not os.path.exists(ULTIMO_REPORTE_MENSUAL_PATH):
        # Primera vez que corre el programa: se guarda el mes ACTUAL como
        # punto de partida, sin generar nada todavía (apenas estamos dentro
        # de ese mes). El primer reporte real saldrá en cuanto este mes
        # termine.
        with open(ULTIMO_REPORTE_MENSUAL_PATH, "w") as f:
            f.write(id_mes_actual)
        return

    with open(ULTIMO_REPORTE_MENSUAL_PATH, "r", encoding="utf-8-sig") as f:
        id_mes_guardado = f.read().strip()

    if id_mes_guardado == id_mes_actual:
        return  # seguimos dentro del mismo mes, nada que generar todavía

    # Cambiamos de mes: se genera el reporte del que ACABA DE TERMINAR
    # (id_mes_guardado), no del que apenas empieza hoy.
    try:
        inicio_anterior, _ = rango_desde_identificador_mes(id_mes_guardado)
        etiqueta_periodo = inicio_anterior.strftime('%B %Y')
        # Siempre se genera primero en la carpeta LOCAL (respaldo garantizado).
        ruta_generada = generar_reporte_mensual_detallado(
            DB_PATH, CARPETA_REPORTES_LOCAL, id_mes_guardado, etiqueta_periodo
        )
        if ruta_generada:
            logging.info(f"Reporte mensual detallado generado ({etiqueta_periodo}): {ruta_generada}")
            distribuir_reporte(ruta_generada)
            enviar_reporte_por_correo(ruta_generada, etiqueta_periodo, tipo="mensual")
        with open(ULTIMO_REPORTE_MENSUAL_PATH, "w") as f:
            f.write(id_mes_actual)
    except Exception:
        logging.exception("Error al generar el reporte mensual detallado")


# ------------------------------------------------------------
# Libro de historial_total.xlsx cacheado en memoria durante toda
# la sesión, para no releer/parsear el archivo completo desde disco
# en cada escaneo (que es lo que más tardaba con el archivo creciendo
# mes con mes). Se sigue GUARDANDO en disco tras cada marca de
# asistencia, para no perder información ante un apagón o cierre
# inesperado; lo que se evita es la RELECTURA repetida.
# ------------------------------------------------------------
_wb_total = None


def obtener_wb_total():
    global _wb_total
    if _wb_total is None:
        _wb_total = load_workbook(HISTORIAL_TOTAL_PATH)
    return _wb_total


def add_to_level_sheet(student_id, nombre, apellido, grado):
    """Marca la asistencia del día en la hoja del grado/categoría correspondiente
    dentro de historial_total.xlsx. Si la hoja no existiera (caso raro, por
    ejemplo si se agregó a alguien a mano sin sincronizar), la crea al vuelo."""
    grado = str(grado)
    wb = obtener_wb_total()
    hoja = wb[grado] if grado in wb.sheetnames else crear_hoja_grado(wb, grado)

    today = datetime.now().strftime("%d-%m")
    col_fecha = None
    for col in range(4, hoja.max_column + 1):
        if hoja.cell(row=1, column=col).value == today:
            col_fecha = col
            break
    if col_fecha is None:
        col_fecha = hoja.max_column + 1
        hoja.cell(row=1, column=col_fecha, value=today)

    fila_persona = None
    for row in range(3, hoja.max_row + 1):
        if hoja.cell(row=row, column=3).value == student_id:
            fila_persona = row
            break
    if fila_persona is None:
        fila_persona = hoja.max_row + 1
        hoja.cell(row=fila_persona, column=1, value=nombre)
        hoja.cell(row=fila_persona, column=2, value=apellido)
        hoja.cell(row=fila_persona, column=3, value=student_id)

    hoja.cell(row=fila_persona, column=col_fecha, value="X")
    wb.save(HISTORIAL_TOTAL_PATH)


def save_statistics(new_df):
    file_name = HISTORIAL_TOTAL_PATH
    if not os.path.exists(file_name):
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            new_df.to_excel(writer, sheet_name="Estadisticas", index=False)
        return
    book = load_workbook(file_name)
    if "Estadisticas" not in book.sheetnames:
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
            new_df.to_excel(writer, sheet_name="Estadisticas", index=False)
        return
    df_existing = pd.read_excel(file_name, sheet_name="Estadisticas")
    df_updated = pd.concat([df_existing, new_df], ignore_index=True)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_updated.to_excel(writer, sheet_name="Estadisticas", index=False)
    # Cualquier cambio hecho por fuera (guardado con pd.ExcelWriter) invalida
    # el libro cacheado en memoria; se recargará solo la próxima vez que se use.
    global _wb_total
    _wb_total = None


# Sincronizar, cerrar el día anterior (si corresponde) y archivar/limpiar
# el mes anterior (si corresponde), ANTES de abrir la interfaz gráfica.
# El orden importa: primero se cierra el día (así, si el cambio de fecha
# también cruza de mes, esa última fila de estadísticas queda dentro del
# archivo del mes que terminó) y luego se revisa el mes. Ninguna de las
# tres cosas debe impedir que el programa abra: si falla, se registra en
# el log y se continúa.
try:
    sync_historial_total()
except Exception:
    logging.exception("Error al sincronizar historial_total.xlsx")

try:
    gestionar_reinicio_diario_automatico()
except Exception:
    logging.exception("Error durante el reinicio diario automático")

try:
    gestionar_limpieza_mensual()
except Exception:
    logging.exception("Error durante la limpieza/archivado mensual")

try:
    gestionar_reporte_quincenal()
except Exception:
    logging.exception("Error durante la revisión del reporte quincenal")

try:
    gestionar_reporte_mensual()
except Exception:
    logging.exception("Error durante la revisión del reporte mensual detallado")


# ============================================================
# INTERFAZ GRÁFICA
# ============================================================

root = tk.Tk()
root.title("Verificación de entrada")
root.geometry("{}x{}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))
root.resizable(False, False)

# Ícono de la ventana (esquina superior izquierda y barra de tareas mientras
# está abierta) -- si no existe 'recursos/logo.ico', simplemente se usa el
# ícono por default de Tkinter, sin error ni mensaje molesto.
try:
    root.iconbitmap(ruta_recurso('logo.ico'))
except Exception:
    pass

root.attributes('-topmost', True)
root.attributes('-fullscreen', True)
root.focus_force()


def cerrar_programa():
    liberar_bloqueo()
    cerrar_conexion_db()
    root.destroy()


root.protocol("WM_DELETE_WINDOW", cerrar_programa)

# Barra superior
barra = tk.Canvas(root, height=36, bg="blue")
barra.pack(fill=tk.X)
barra.grid(row=0, column=0, columnspan=2, sticky="news")

with open(CIERRE_PATH, "r", encoding="utf-8-sig") as file:
    close_pass = [line.strip() for line in file]

with open(REIN_PATH, "r", encoding="utf-8-sig") as file:
    rein_pass = [line.strip().replace("\r", "") for line in file if line.strip()]

with open(SALIR_PANTALLA_PATH, "r", encoding="utf-8-sig") as file:
    salir_pantalla_pass = [line.strip() for line in file if line.strip()]


# ------------------------------------------------------------
# Esc + contraseña: permite salir/entrar del modo pantalla completa sin
# usar la clave de cierre (que apaga todo el programa). Útil para que el
# operador revise algo rápido en Windows sin cerrar la aplicación.
# ------------------------------------------------------------
def manejar_escape(event=None):
    intento = simpledialog.askstring(
        "Salir de pantalla completa",
        "Contraseña:",
        show="*",
        parent=root,
    )
    if intento is None:
        return  # se canceló el cuadro de diálogo
    if intento.strip() in salir_pantalla_pass:
        nuevo_estado = not bool(root.attributes('-fullscreen'))
        root.attributes('-fullscreen', nuevo_estado)
        root.attributes('-topmost', nuevo_estado)
        logging.info("Se cambió el modo de pantalla completa con la contraseña de Esc.")
    else:
        logging.warning("Intento fallido de salir de pantalla completa (contraseña incorrecta).")


root.bind("<Escape>", manejar_escape)


import tkinter.font as tkfont

# ---- Estado de las TRES categorías, cargadas simultáneamente ----
# Ya no existe una "categoría activa": como todos pasan por el mismo
# lector, cada ID escaneado se busca en las tres listas de IDs válidos y
# se clasifica solo — sin que el operador tenga que elegir nada primero.
estado_categorias = {}  # idx -> {valid_ids, valid_ids_set, historial, historial_set, num_strings, num_actual}

# Evita que un doble disparo del lector de código de barras (o un doble
# Enter accidental) procese dos veces la misma entrada mientras la primera
# todavía se está guardando en disco.
procesando_entrada = False


def cargar_todas_categorias():
    """(Re)carga desde disco los IDs válidos y el historial diario de las
    TRES categorías a la vez. Se llama al iniciar, y de nuevo cada vez que
    el conteo se reinicia (manual o automáticamente)."""
    global estado_categorias
    nuevo_estado = {}
    for idx, cat in enumerate(CATEGORIAS):
        valid_ids = leer_ids_categoria(cat["ids_path"])
        valid_ids_set = set(valid_ids['ID'])

        historial = pd.read_excel(cat["historial_path"], dtype={'ID': str})
        historial['ID'] = historial['ID'].astype(str).str.strip()
        historial_set = set(historial['ID'])

        nuevo_estado[idx] = {
            "valid_ids": valid_ids,
            "valid_ids_set": valid_ids_set,
            "historial": historial,
            "historial_set": historial_set,
            "num_strings": valid_ids['ID'].count(),
            "num_actual": historial['ID'].count(),
        }
    estado_categorias = nuevo_estado


def buscar_categoria_de_id(student_id):
    """Busca en las tres categorías cuál tiene este ID registrado.
    Devuelve el índice de esa categoría, o None si no aparece en
    ninguna (ID inválido)."""
    for idx in range(len(CATEGORIAS)):
        if student_id in estado_categorias[idx]["valid_ids_set"]:
            return idx
    return None


cargar_todas_categorias()

# Etiqueta de estado (a la izquierda de la barra)
status_id = barra.create_text(
    5, 20,
    text="Estado: Standby",
    font=("Arial", 16, "bold"),
    fill="white",
    anchor="w"
)

# Contadores de las tres categorías, visibles todos al mismo tiempo en la
# barra superior (a la derecha), sin necesidad de cambiar entre ellos.
FUENTE_CONTADOR = ("Arial", 12, "bold")
_fuente_medida_contador = tkfont.Font(family="Arial", size=12, weight="bold")
_GAP_CONTADORES = 25
_MARGEN_DERECHO = 15

contador_ids = {}


def texto_contador(idx):
    cat = CATEGORIAS[idx]
    est = estado_categorias[idx]
    return f"{cat['nombre']}: {est['num_actual']}/{est['num_strings']}"


def actualizar_contadores():
    """Dibuja/actualiza los tres contadores, alineados por la derecha y
    calculando su ancho real (con la fuente ya cargada) para que nunca se
    encimen entre sí, sin importar cuánto crezcan los números."""
    x = root.winfo_screenwidth() - _MARGEN_DERECHO
    for idx in reversed(range(len(CATEGORIAS))):
        texto = texto_contador(idx)
        if idx not in contador_ids:
            contador_ids[idx] = barra.create_text(
                x, 20, text=texto, font=FUENTE_CONTADOR, fill="white", anchor="e"
            )
        else:
            barra.coords(contador_ids[idx], x, 20)
            barra.itemconfig(contador_ids[idx], text=texto)
        x -= _fuente_medida_contador.measure(texto) + _GAP_CONTADORES


actualizar_contadores()


# ------------------------------------------------------------
# Colores de estado centralizados: cada evento cambia tanto la barra
# superior como el fondo/texto de la etiqueta grande de estado, para que
# el resultado se note sin tener que mirar justo la barra de arriba (y sin
# depender solo del color, ya que también cambia el texto/contraste).
# ------------------------------------------------------------
COLORES_ESTADO = {
    "standby":   {"barra": "blue",  "status_bg": "#eaf2ff", "status_fg": "#0b3d91"},
    "permitido": {"barra": "green", "status_bg": "#eaffea", "status_fg": "#0b6e0b"},
    "alerta":    {"barra": "red",   "status_bg": "#ffecec", "status_fg": "#a30000"},
}


def aplicar_estado_visual(estado):
    colores = COLORES_ESTADO[estado]
    barra.configure(bg=colores["barra"])
    status_label.configure(bg=colores["status_bg"], fg=colores["status_fg"])


def mostrar_confirmacion_temporal(mensaje, ms=2200, color_bg="#0b6e0b"):
    """Ventana emergente grande que se autodestruye sola, para confirmar
    visualmente una acción rápida (como el reinicio del conteo) que de otro
    modo podría pasar desapercibida si el operador escanea muy rápido."""
    ventana = tk.Toplevel(root)
    ventana.overrideredirect(True)
    ventana.attributes('-topmost', True)
    ancho, alto = 560, 140
    x = (root.winfo_screenwidth() - ancho) // 2
    y = (root.winfo_screenheight() - alto) // 2
    ventana.geometry(f"{ancho}x{alto}+{x}+{y}")
    ventana.configure(bg=color_bg)
    tk.Label(
        ventana, text=mensaje, font=("Arial", 20, "bold"),
        bg=color_bg, fg="white", wraplength=520, justify="center"
    ).pack(expand=True, fill="both")
    ventana.after(ms, ventana.destroy)


# Función para verificar entrada
def verify_entry():
    global procesando_entrada

    if procesando_entrada:
        # Ya hay un escaneo en proceso; se ignora este disparo repetido.
        return
    procesando_entrada = True

    try:
        try:
            hubo_reinicio_automatico = gestionar_reinicio_diario_automatico()
        except Exception:
            logging.exception("Error al revisar el reinicio diario automático durante un escaneo")
            hubo_reinicio_automatico = False

        if hubo_reinicio_automatico:
            try:
                gestionar_limpieza_mensual()
            except Exception:
                logging.exception("Error al revisar la limpieza mensual durante un escaneo")

            try:
                gestionar_reporte_quincenal()
            except Exception:
                logging.exception("Error al revisar el reporte quincenal durante un escaneo")

            try:
                gestionar_reporte_mensual()
            except Exception:
                logging.exception("Error al revisar el reporte mensual detallado durante un escaneo")

            cargar_todas_categorias()
            actualizar_contadores()
            aplicar_estado_visual("standby")
            barra.itemconfig(status_id, text="Estado: Standby", fill="white")
            mostrar_confirmacion_temporal("Nuevo día detectado: el conteo se reinició automáticamente.")

        student_id = text_id.get("1.0", tk.END).strip()
        student_id = str(student_id)

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        fecha = datetime.now().strftime("%Y-%m-%d")
        hora = datetime.now().strftime("%H:%M:%S")

        nombre = apellido = grado = None

        if not student_id:
            status_label.config(text="Ninguna ID escaneada todavia.")
            barra.itemconfig(status_id, text="Estado: Standby", fill="white")
            aplicar_estado_visual("standby")
            return

        # Clasificación automática: se busca el ID en las tres categorías,
        # sin que el operador tenga que elegir cuál está activa.
        idx_categoria = buscar_categoria_de_id(student_id)

        if idx_categoria is not None:
            cat = CATEGORIAS[idx_categoria]
            est = estado_categorias[idx_categoria]

            row_index = est["valid_ids"].index[est["valid_ids"]['ID'] == student_id].tolist()
            nombre = est["valid_ids"].loc[row_index[0], 'Nombre']
            apellido = est["valid_ids"].loc[row_index[0], 'Apellido']
            grado = est["valid_ids"].loc[row_index[0], 'Grado']

            if student_id in est["historial_set"]:
                status_label.config(text=f"{student_id} - {nombre} {apellido} {grado} ({cat['nombre']}) ya ha entrado hoy.")
                barra.itemconfig(status_id, text="Estado: Ya ingresado", fill="white")
                aplicar_estado_visual("alerta")
            else:
                status_label.config(text=f"{student_id} - {nombre} {apellido} {grado} ({cat['nombre']}) permitido")
                barra.itemconfig(status_id, text="Estado: Permitido", fill="white")
                aplicar_estado_visual("permitido")
        else:
            status_label.config(text=f"ID inválido: {student_id}")
            barra.itemconfig(status_id, text="Estado: ID inválido", fill="white")
            aplicar_estado_visual("alerta")

        if student_id in close_pass:
            cerrar_programa()
            return

        if student_id in rein_pass:
            ejecutar_reinicio_diario(motivo="código manual (rein.txt)")
            cargar_todas_categorias()

            status_label.config(text="El recuento de entradas ha sido restablecido.")
            barra.itemconfig(status_id, text="Estado: Standby", fill="white")
            aplicar_estado_visual("standby")
            actualizar_contadores()
            mostrar_confirmacion_temporal("Recuento de entradas reiniciado correctamente.")
            return

        # agregar entrada al historial de eventos en pantalla, con color según el resultado
        if idx_categoria is None:
            status = "ID inválido"
            history_text.insert(tk.END, f"{now} - ID: {student_id} - {status}\n", "error")
            reproducir_sonido_error()
        elif student_id in estado_categorias[idx_categoria]["historial_set"]:
            status = "Ingreso no permitido, ya entró hoy"
            history_text.insert(tk.END, f"{now} - ID: {student_id} ({nombre} {apellido} {grado}, {CATEGORIAS[idx_categoria]['nombre']}) - {status}\n", "warn")
            reproducir_sonido_error()
        else:
            status = "Ingreso permitido"
            history_text.insert(tk.END, f"{now} - ID: {student_id} ({nombre} {apellido} {grado}, {CATEGORIAS[idx_categoria]['nombre']}) - {status}\n", "ok")
        history_text.see(tk.END)

        # Añadir al historial diario de la categoría detectada (en memoria + disco)
        if idx_categoria is not None and student_id not in estado_categorias[idx_categoria]["historial_set"]:
            cat = CATEGORIAS[idx_categoria]
            est = estado_categorias[idx_categoria]
            new_data = {
                'Fecha': [fecha], 'Hora': [hora], 'ID': [student_id],
                'Nombre': [nombre], 'Apellido': [apellido], 'Grado': [grado],
            }
            new_df = pd.DataFrame(new_data)
            est["historial"] = pd.concat([est["historial"], new_df], ignore_index=True)
            est["historial"].to_excel(cat["historial_path"], index=False)
            est["historial_set"].add(student_id)
            est["num_actual"] = est["historial"]['ID'].count()
            add_to_level_sheet(student_id, nombre, apellido, grado)
            registrar_entrada_db(student_id, fecha, hora)

        actualizar_contadores()

    except Exception:
        logging.exception("Error inesperado al procesar el ID escaneado")
        status_label.config(text="Ocurrió un error inesperado. Revisa app.log.")
        barra.itemconfig(status_id, text="Estado: Error", fill="white")
        aplicar_estado_visual("alerta")

    finally:
        procesando_entrada = False
        try:
            text_id.delete("1.0", tk.END)
            text_id.focus_set()
        except tk.TclError:
            pass  # la ventana ya se cerró (ej. se escaneó la clave de cierre)


# etiqueta para mostrar la identificacion escaneada
id_label = tk.Label(root, text="ID escaneada:", font=("Arial", 16), fg="grey")
id_label.grid(row=2, column=0, columnspan=2, sticky="ew")

# entrada de texto para id
text_id = tk.Text(root, width=20, height=1, font=("Arial", 14))
text_id.grid(row=3, column=0, columnspan=2, sticky="ew", padx=300)

# etiqueta de estado para respuesta a tiempo real (colores iniciales = standby)
status_label = tk.Label(root, text="Esperando ID", font=("Arial", 20))
status_label.grid(row=1, column=0, columnspan=2, sticky="ew")
aplicar_estado_visual("standby")

# historial de entradas, con colores según el resultado de cada escaneo
history_text = tk.Text(root, width=40, height=10, font=("Arial", 12), wrap=tk.WORD)
history_text.grid(row=4, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)
history_text.tag_configure("ok", foreground="#0b6e0b")
history_text.tag_configure("warn", foreground="#b36b00")
history_text.tag_configure("error", foreground="#a30000")

# activa la verificacion tras el escaneo con el lector de codigo de barras
text_id.bind("<KeyRelease-Return>", lambda event: verify_entry())

# selecciona la entrada de texto automaticamente
text_id.focus_set()

# Crédito discreto: más pequeño y tenue para no competir visualmente con la
# información que sí necesita ver el operador.
creat = tk.Label(root, text="Aplicación desarrollada estudiantes de 11C — IELP PROM 2026", font=("Arial", 9), fg="#b5b5b5")
creat.grid(row=5, column=0, columnspan=2, sticky="ew", pady=6)

root.grid_rowconfigure(0, weight=0)
root.grid_rowconfigure(1, weight=2)
root.grid_rowconfigure(2, weight=0)
root.grid_rowconfigure(3, weight=1)
root.grid_rowconfigure(4, weight=2)
root.grid_rowconfigure(5, weight=0)
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=1)

# ------------------------------------------------------------
# Reloj en la esquina inferior derecha. Se coloca con place() (no con
# grid()) para que quede fijo en esa esquina de la ventana sin importar
# cómo se acomoden las demás filas/columnas, y se actualiza solo cada
# segundo con root.after (sin bloquear el resto de la interfaz).
# ------------------------------------------------------------
reloj_label = tk.Label(root, font=("Arial", 13, "bold"), fg="#555555")
reloj_label.place(relx=1.0, rely=1.0, anchor="se", x=-14, y=-8)


def actualizar_reloj():
    reloj_label.config(text=datetime.now().strftime("%H:%M:%S   %d/%m/%Y"))
    root.after(1000, actualizar_reloj)


actualizar_reloj()

root.mainloop()
